# -*- coding: utf-8 -*-
"""
Created on Mon Nov 27 07:57:53 2017

@author: Raki
"""

import numpy as np
from openpyxl import load_workbook
import random
import time

class BatClustering:
    def __init__(self, d, nbat, k, ngen, a, r, qmin, qmax, function):
        self.d = d
        self.nbat = nbat
        self.k = k
        self.Fun = function
        
        self.ngen = ngen  #generations
        self.a = a  #loudness
        self.r = r  #pulse rate
        self.qmin = qmin  #frequency min
        self.qmax = qmax  #frequency max

        self.f_min = 0.0  #minimum fitness
        
        self.q = np.zeros((self.nbat))  #frequency

        self.v = np.zeros((self.nbat, self.k, self.d))  #velocity
        self.sol = np.zeros((self.nbat, self.k, self.d))  #population of solutions
        self.fitness = np.zeros((self.nbat))  #fitness
        self.best = np.zeros((self.k, self.d))  #best solution
        
    def clustering(self, data):
        bound = np.zeros((self.d, 2))
        for i in range(0, self.d):
            bound[i, 0] = data[np.argmin(data[:, i]), i]
            bound[i, 1] = data[np.argmax(data[:, i]), i]
            
        s = np.zeros((self.nbat, self.k, self.d))
            
        self.init_bat(data)
        for i in range(self.nbat):
            self.sol[i] = self.simplebound(self.sol[i], bound)
        
        x = 0 
        for t in range(self.ngen):
            for i in range(self.nbat):
                x += 1
                rnd = random.random()
                
                self.q[i] = self.qmin + (self.qmin-self.qmax)*rnd
                self.v[i] += (self.sol[i]-self.best)*self.q[i]
                s[i] = self.sol[i] + self.v[i]
                
                s[i] = self.simplebound(s[i], bound)
                
                rnd = random.random()
                
                fnew, z = self.Fun(s[i], self.k, data)
                
                if(rnd>self.r):
                    s[i] = self.simplebound(s[i], bound)
                
                rnd = random.random()
                
                if(fnew <= self.fitness[i]) and (rnd < self.a):
                    self.sol[i] = s[i]
                    self.fitness[i] = fnew
                if(fnew <= self.f_min):
                    self.best = s[i]
                    self.f_min = fnew
#            x += 1
#            rnd = np.r
#            self.q = self.qmin + (self.)
                print(x, ". Fitness terbaik : ", self.f_min)
        bestfitness, clusters = self.Fun(self.best, self.k, data)
        return(self.best, bestfitness, clusters)
    
    def simplebound(self, val, bound):
        for i in range(0, self.k):
            for j in range(0, self.d):
                if(val[i, j]<bound[j, 0]):
                    val[i, j] = bound[j, 0]
                elif(val[i, j]>bound[j, 1]):
                    val[i, j] = bound[j, 1]
        return val
    
    def best_bat(self):
        i = 0
        j = 0
        for i in range(self.nbat):
            if (self.fitness[i] < self.fitness[j]):
                j = i
        for x in range(self.k):
            for i in range(self.d):
                self.best[x, i] = self.sol[j, x, i]
        self.f_min = self.fitness[j]

    def init_bat(self, data):
        for i in range(self.nbat):
            self.sol[i, :] = self.initcentroid(data)
            self.fitness[i], z = self.Fun(self.sol[i], self.k, data)
        self.best_bat()
    
    def initcentroid(self, data):
        centroid = np.zeros((self.k, self.d))
        centroid[0, :] = data[int(random.random()*len(data)), :]
        for i in range(0, self.k-1):
            centroid[i+1, :] = data[int(random.random()*len(data)), :]
        return centroid
    
    def euclideandist(self, x, y):
        sumof = np.zeros((len(x), 1))
        for i in range(0, len(y)):
            sumof[:, 0] += np.power(np.absolute(x[:, i]-y[i]), 2)
        sumof = np.power(sumof, 1/2)
        return sumof[:, 0]
    
def dist(x, y):
    sumof = np.zeros((len(x), 1))
    for i in range(0, len(y)):
        sumof[:, 0] += np.absolute(x[:, i]-y[i])**2
    return np.sqrt(sumof[:, 0])
    
def Fun(sol, k, data):
    row_count = len(data)
    
    distance = np.empty((row_count, k))
    datacentroid = np.zeros((row_count), dtype=np.int32)
        
    for i in range(k):
        distance[:, i] = dist(data, sol[i])
        
    for i in range(row_count):
        datacentroid[i] = np.argmin(distance[i, :])
        
    clusters = np.zeros((k), dtype=np.int16)
    for i in range(row_count):
        clusters[int(datacentroid[i])] += 1
        
    x = 0
    distfromc = np.zeros((row_count))
    for i in range(k):
        datakidx = np.argwhere(datacentroid==i)
        rowi = len(datakidx)
        datak = np.empty((rowi, len(sol[0])))
        datak[:] = data[datakidx[:, 0]]
        if(rowi>0):
            distfromc[x:x+rowi] = dist(datak, sol[i])
            x += rowi
            
    return np.sum(distfromc)/k, clusters
        
if __name__=="__main__":
    np.set_printoptions(threshold=np.nan)
    fname = 'Dataset-Tugas-3-SI.xlsx'
    wb = load_workbook(filename = fname)
    
    sheet_train = wb['data']
    
    col_count = sheet_train.max_column
    row_count = sheet_train.max_row
    
    data = np.empty((row_count, col_count))
    
    for i in range(0, col_count):
        data[:, i] = np.array([r[i].value for r in sheet_train.iter_rows()])
    
    start_time = time.time()
    
    bc = BatClustering(col_count, 20, 10, 5000, 0.8, 0.3, 0, 0.4, Fun)
    [centroid, fitness, cluster] = bc.clustering(data)
    
    print(cluster)
    print(centroid)
    print(fitness)
    
    print("--- %s seconds ---" % (time.time() - start_time))